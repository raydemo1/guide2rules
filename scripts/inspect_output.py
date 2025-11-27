import os
import openpyxl

def main():
    root = os.path.dirname(os.path.dirname(__file__))
    path = os.path.join(root, "outputs", "transportation", "sample.classified.xlsx")
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    print("HEADERS:", headers)
    rows = []
    for row in ws.iter_rows(min_row=2, max_row=3):
        rows.append([c.value for c in row])
    print("ROWS:")
    for r in rows:
        print(r)

if __name__ == "__main__":
    main()
