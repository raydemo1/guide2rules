import csv
import json
import os
from typing import Dict, List, Any
import pandas as pd
from data_preprocess import _normalize_df, _canonical_row, _compose_text


def read_json(path: str) -> Any:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def list_extractions(domain_dir: str) -> List[str]:
    files = []
    for f in os.listdir(domain_dir):
        if f.endswith(".extraction.json") or f.endswith(".extraction.detailed.json") or f.endswith(".items.normalized.json"):
            files.append(os.path.join(domain_dir, f))
    return files


def collect_static_rules(extraction: Dict) -> List[Dict]:
    items = extraction.get("extraction", [])
    print(f"[DEBUG] Layer3 收集静态规则，抽取项数量: {len(items)}")
    rows: List[Dict] = []
    idx = 0
    for it in items:
        # 支持分组结构：items 列表 + 共享 citation
        if isinstance(it.get("items"), list):
            p = it.get("path", {})
            citation = it.get("citation", {})
            category = "/".join([p.get("level1",""), p.get("level2",""), p.get("level3",""), p.get("level4","")]).strip("/")
            for sub in it["items"]:
                idx += 1
                field = sub.get("name") or sub.get("canonical", "")
                row = {
                    "RuleId": f"S-{idx}",
                    "FieldName": field,
                    "Category": category,
                    "Level": sub.get("level", ""),
                    "ConditionExpression": ";".join(sub.get("conditions", []) or []),
                    "ExceptionExpression": ";".join(sub.get("exceptions", []) or []),
                    "Citation": json.dumps(citation, ensure_ascii=False),
                    "Priority": 50,
                    "Source": extraction.get("source", "")
                }
                rows.append(row)
        else:
            idx += 1
            field = it.get("item", {}).get("name") or it.get("item", {}).get("canonical") or it.get("field", "")
            if it.get("path"):
                p = it.get("path")
                category = "/".join([p.get("level1",""), p.get("level2",""), p.get("level3",""), p.get("level4","")]).strip("/")
            else:
                category = it.get("category", "")
            row = {
                "RuleId": f"S-{idx}",
                "FieldName": field,
                "Category": category,
                "Level": it.get("level", ""),
                "ConditionExpression": ";".join(it.get("conditions", []) or []),
                "ExceptionExpression": ";".join(it.get("exceptions", []) or []),
                "Citation": json.dumps(it.get("citation", {}), ensure_ascii=False),
                "Priority": 50,
                "Source": extraction.get("source", "")
            }
            rows.append(row)
    print(f"[DEBUG] 收集到 {len(rows)} 条静态规则")
    return rows


def try_read_xlsx(path: str) -> List[Dict]:
    try:
        import openpyxl
    except Exception:
        return []
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    out: List[Dict] = []
    for row in ws.iter_rows(min_row=2):
        values = [c.value for c in row]
        out.append({h: ("" if v is None else str(v)) for h, v in zip(headers, values)})
    return out


def try_read_csv(path: str) -> List[Dict]:
    with open(path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        return [{k: (v or "") for k, v in row.items()} for row in reader]


def read_dynamic_rules(domain: str, root: str) -> List[Dict]:
    dir_path = os.path.join(root, "excels", domain)
    print(f"[DEBUG] 检查动态规则目录: {dir_path}")

    if not os.path.isdir(dir_path):
        print(f"[DEBUG] 目录不存在，跳过动态规则读取")
        return []

    files = os.listdir(dir_path)
    print(f"[DEBUG] 找到 {len(files)} 个文件: {files}")

    rows: List[Dict] = []
    for f in files:
        p = os.path.join(dir_path, f)
        if f.lower().endswith(".csv"):
            print(f"[DEBUG] 读取CSV文件: {f}")
            csv_rows = try_read_csv(p)
            rows.extend(csv_rows)
            print(f"[DEBUG] CSV文件包含 {len(csv_rows)} 行数据")
        elif f.lower().endswith(".xlsx"):
            print(f"[DEBUG] 读取Excel文件: {f}")
            xlsx_rows = try_read_xlsx(p)
            rows.extend(xlsx_rows)
            print(f"[DEBUG] Excel文件包含 {len(xlsx_rows)} 行数据")

    print(f"[DEBUG] 原始动态规则总数: {len(rows)}")

    normalized: List[Dict] = []
    for i, r in enumerate(rows, 1):
        normalized.append({
            "RuleId": f"D-{i}",
            "FieldName": r.get("FieldName", r.get("field", "")),
            "Category": r.get("Category", r.get("category", "")),
            "Level": r.get("Level", r.get("level", "")),
            "ConditionExpression": r.get("Pattern", r.get("condition", "")),
            "ExceptionExpression": r.get("Exception", r.get("exception", "")),
            "Citation": r.get("Rationale", r.get("citation", "")),
            "Priority": int(r.get("Priority", r.get("priority", 60)) or 60),
            "Source": f
        })

    print(f"[DEBUG] 规范化后的动态规则数量: {len(normalized)}")
    return normalized


def read_dynamic_rules_excel(domain: str, root: str) -> List[Dict]:
    dir_path = os.path.join(root, "excels", domain)
    if not os.path.isdir(dir_path):
        return []
    out: List[Dict] = []
    for base, _, files in os.walk(dir_path):
        for f in files:
            if not f.lower().endswith(".xlsx"):
                continue
            p = os.path.join(base, f)
            try:
                sheets = pd.read_excel(p, sheet_name=None)
            except Exception:
                continue
            for sname, df in sheets.items():
                if df is None or df.empty:
                    continue
                df = _normalize_df(df)
                cols = set(df.columns)
                if not ("fieldname" in cols or "fielddesc" in cols or "classification" in cols):
                    continue
                for _, r in df.iterrows():
                    x = _canonical_row(r)
                    if not x.get("fieldname") and not x.get("fielddesc"):
                        continue
                    if not x.get("classification"):
                        continue
                    s = str(x.get("grade", "")).strip().upper().replace("（", "(").replace("）", ")")
                    sn = s.replace(" ", "")
                    if "S1" in sn or "一级" in s or "G1" in sn or "G-1" in sn:
                        lvl = "S1"
                    elif "S2" in sn or "二级" in s or "G2" in sn or "G-2" in sn:
                        lvl = "S2"
                    elif "S3" in sn or "三级" in s or "G3" in sn or "G-3" in sn:
                        lvl = "S3"
                    elif "S4" in sn or "四级" in s or "G4" in sn or "G-4" in sn:
                        lvl = "S4"
                    else:
                        lvl = ""
                    citation = _compose_text(x)
                    category = x.get("classification")
                    out.append({
                        "RuleId": f"D-XLSX-{len(out)+1}",
                        "FieldName": x.get("fieldname"),
                        "Category": category,
                        "Level": lvl,
                        "ConditionExpression": "",
                        "ExceptionExpression": "",
                        "Citation": citation,
                        "Priority": 60,
                        "Source": p
                    })
    return out


def write_openrules_table(domain: str, root: str, rows: List[Dict]):
    print(f"[DEBUG] 写入OpenRules表格，域名: {domain}，规则数量: {len(rows)}")

    out_dir = os.path.join(root, "openrules", domain)
    os.makedirs(out_dir, exist_ok=True)
    print(f"[DEBUG] 输出目录: {out_dir}")

    csv_path = os.path.join(out_dir, "Classifier.csv")
    json_path = os.path.join(out_dir, "Classifier.json")

    headers = [
        "RuleId",
        "FieldName",
        "Category",
        "Level",
        "ConditionExpression",
        "ExceptionExpression",
        "Citation",
        "Priority",
        "Source"
    ]

    print(f"[DEBUG] 写入CSV文件: {csv_path}")
    with open(csv_path, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        for r in rows:
            writer.writerow({k: r.get(k, "") for k in headers})

    print(f"[DEBUG] 写入JSON文件: {json_path}")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump({"domain": domain, "rows": rows}, f, ensure_ascii=False, indent=2)

    print(f"[DEBUG] OpenRules表格写入完成: {csv_path}")
    return csv_path


def _norm(s: Any) -> str:
    if s is None:
        return ""
    return str(s).strip()


def _level_rank(level: str) -> int:
    m = {"S1": 1, "S2": 2, "S3": 3, "S4": 4}
    return m.get(_norm(level).upper(), 0)


def resolve_conflicts(rows: List[Dict]) -> List[Dict]:
    print(f"[DEBUG] 开始去重与冲突消解，输入 {len(rows)} 条")
    buckets: Dict[str, Dict] = {}
    for r in rows:
        k1 = _norm(r.get("FieldName")).lower()
        k2 = _norm(r.get("Category")).lower()
        if not k1 and not k2:
            continue
        key = k1 + "|" + k2
        cur = buckets.get(key)
        if not cur:
            buckets[key] = r
            continue
        a = cur
        b = r
        ra = _level_rank(a.get("Level"))
        rb = _level_rank(b.get("Level"))
        chosen = a
        other = b
        if rb > ra:
            chosen = b
            other = a
        elif rb == ra:
            try:
                pa = int(a.get("Priority", 0))
            except Exception:
                pa = 0
            try:
                pb = int(b.get("Priority", 0))
            except Exception:
                pb = 0
            if pb < pa:
                chosen = b
                other = a
        conds = []
        for s in [_norm(chosen.get("ConditionExpression")), _norm(other.get("ConditionExpression"))]:
            if s:
                conds.append(s)
        exs = []
        for s in [_norm(chosen.get("ExceptionExpression")), _norm(other.get("ExceptionExpression"))]:
            if s:
                exs.append(s)
        cits = []
        for s in [_norm(chosen.get("Citation")), _norm(other.get("Citation"))]:
            if s:
                cits.append(s)
        srcs = []
        for s in [_norm(chosen.get("Source")), _norm(other.get("Source"))]:
            if s:
                srcs.append(s)
        chosen["ConditionExpression"] = " OR ".join(sorted(set(conds))) if conds else ""
        chosen["ExceptionExpression"] = " OR ".join(sorted(set(exs))) if exs else ""
        chosen["Citation"] = " || ".join(sorted(set(cits))) if cits else ""
        chosen["Source"] = ";".join(sorted(set(srcs)))
        buckets[key] = chosen
    merged = list(buckets.values())
    for i, r in enumerate(merged, 1):
        r["RuleId"] = f"R-{i}"
    print(f"[DEBUG] 去重后剩余 {len(merged)} 条")
    return merged


def main():
    print(f"[DEBUG] Layer3 主程序开始执行")

    root = os.path.dirname(os.path.dirname(__file__))
    artifacts = os.path.join(root, "artifacts")
    excels = os.path.join(root, "excels")
    print(f"[DEBUG] 工件目录: {artifacts}")
    print(f"[DEBUG] Excel目录: {excels}")

    domains = set()
    if os.path.isdir(artifacts):
        artifact_domains = [d for d in os.listdir(artifacts) if os.path.isdir(os.path.join(artifacts, d))]
        domains.update(artifact_domains)
        print(f"[DEBUG] 工件目录中的域: {artifact_domains}")

    if os.path.isdir(excels):
        excel_domains = [d for d in os.listdir(excels) if os.path.isdir(os.path.join(excels, d))]
        domains.update(excel_domains)
        print(f"[DEBUG] Excel目录中的域: {excel_domains}")

    print(f"[DEBUG] 总共找到 {len(domains)} 个域: {sorted(domains)}")

    for domain in sorted(domains):
        print(f"\n[DEBUG] ========== 处理域: {domain} ==========")

        static_rows: List[Dict] = []
        domain_artifacts = os.path.join(artifacts, domain)
        if os.path.isdir(domain_artifacts):
            print(f"[DEBUG] 处理静态规则，工件目录: {domain_artifacts}")
            ex_files = list_extractions(domain_artifacts)
            print(f"[DEBUG] 找到 {len(ex_files)} 个抽取文件: {ex_files}")

            for p in ex_files:
                print(f"[DEBUG] 处理抽取文件: {p}")
                static_rows.extend(collect_static_rules(read_json(p)))

            print(f"[DEBUG] 静态规则总数: {len(static_rows)}")
        else:
            print(f"[DEBUG] 域 {domain} 的工件目录不存在，跳过静态规则")

        # 合并两种来源：CSV/XLSX结构化规则表 + Excel字段表预处理
        print(f"[DEBUG] 读取动态规则...")
        dynamic_rows = read_dynamic_rules(domain, root) + read_dynamic_rules_excel(domain, root)
        print(f"[DEBUG] 动态规则总数: {len(dynamic_rows)}")

        combined = static_rows + dynamic_rows
        print(f"[DEBUG] 合并后总规则数: {len(combined)}")
        combined = resolve_conflicts(combined)

        if not combined:
            print(f"[DEBUG] 域 {domain} 没有规则，跳过")
            continue

        out_path = write_openrules_table(domain, root, combined)
        print(f"[DEBUG] 输出文件: {out_path}")

    print(f"\n[DEBUG] Layer3 主程序执行完成")


if __name__ == "__main__":
    main()
