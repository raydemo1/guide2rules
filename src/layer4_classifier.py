import os
import sys
import json
import argparse
from typing import Dict, List, Any
import openpyxl
from tqdm import tqdm

# 添加项目根目录到路径
sys.path.append(os.path.dirname(os.path.dirname(__file__)))
from business_rules.engine import run_all
from rules.variables import ClassificationVariables
from rules.actions import ClassificationActions


def load_rules(domain: str, root: str) -> List[Dict]:
    path = os.path.join(root, "rules", domain, "unified_rules.json")
    if not os.path.exists(path):
        raise FileNotFoundError(f"unified_rules.json not found for domain {domain}")
    with open(path, "r", encoding="utf-8") as f:
        rules: List[Dict] = json.load(f)
    print(f"[INFO] Loaded unified rules: count={len(rules)}")
    return rules


def detect_columns(headers: List[str]) -> Dict[str, int]:
    hmap = {str(h or "").strip(): i for i, h in enumerate(headers)}

    def find(candidates: List[str]) -> int:
        for k in candidates:
            if k in hmap:
                return hmap[k]
        return -1

    field_en_idx = find(["字段名", "FieldName", "字段名称", "名称", "英文字段名"]) 
    field_cn_idx = find(["字段注释", "中文字段名", "中文名称", "字段中文名"]) 
    value_idx = find(["字段样本", "样本", "数据示例"])  
    table_idx = find(["表名", "Table", "表名称"])  
    return {"field_en": field_en_idx, "field_cn": field_cn_idx, "value": value_idx, "table": table_idx}


def normalize_category(category: str) -> str:
    s = str(category or "").strip()
    s = s.replace("\\", "/")
    s = "/".join([p.strip() for p in s.split("/") if p.strip()])
    return s


def classify_rows(
    domain: str,
    in_path: str,
    out_path: str,
    stop_first: bool,
    sheet_name: str,
):
    root = os.path.dirname(os.path.dirname(__file__))
    try:
        unified_rules = load_rules(domain, root)
        print(f"[DEBUG] Loaded unified rules from {domain}: {len(unified_rules)}")
    except Exception as e:
        print(f"[ERROR] Failed to load rules: {e}")
        return

    # 过滤无效正则的规则（空值或不可编译），避免边缘情况误命中
    def _valid_rule(rule: Dict) -> bool:
        cond = rule.get("conditions") or {}
        def _iter(c):
            if not isinstance(c, dict):
                return []
            items = []
            for k in ("all", "any"):
                arr = c.get(k) or []
                for it in arr:
                    if isinstance(it, dict):
                        items.append(it)
            return items
        stack = [cond]
        while stack:
            cur = stack.pop()
            for it in _iter(cur):
                name = str(it.get("name", ""))
                op = str(it.get("operator", ""))
                val = it.get("value")
                if name == "value_text" and op == "matches_regex":
                    rx = str(val or "")
                    if not rx:
                        return False
                    try:
                        __import("re").compile(rx)
                    except Exception:
                        return False
        return True

    unified_rules = [r for r in unified_rules if _valid_rule(r)]

    wb = openpyxl.load_workbook(in_path)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    headers = [
        str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))
    ]
    cols = detect_columns(headers)

    # 调试打印列索引
    print(f"[DEBUG] Column mapping: {cols}")

    rows = []
    for row in ws.iter_rows(min_row=2):
        rows.append([c.value for c in row])

    processed = []
    total_rows = len(rows)
    matched_rows = 0

    import re
    def make_tokens(s: str) -> str:
        parts = [p for p in re.split(r"[^A-Za-z]+|_+", (s or "").lower()) if p]
        return " ".join(parts)

    for r in tqdm(rows, desc=f"[PROGRESS] {os.path.basename(in_path)}", unit="row"):
        field_en = str((r[cols["field_en"]] if cols.get("field_en", -1) >= 0 else "") or "").strip()
        field_cn = str((r[cols["field_cn"]] if cols.get("field_cn", -1) >= 0 else "") or "").strip()
        field_name = field_en
        field_comment = field_cn or field_en
        table_name = (
            str(r[cols["table"]]) if cols.get("table", -1) >= 0 and r[cols["table"]] is not None else ws.title
        )
        category = ""
        value_text = str(r[cols["value"]]) if cols["value"] >= 0 and r[cols["value"]] is not None else ""

        obj = {
            "field_name": field_name,
            "field_comment": field_comment,
            "table_name": table_name,
            "field_tokens": make_tokens(field_name),
            "table_tokens": make_tokens(table_name),
            "category_path": category,
            "value_text": value_text,
            "score": 0,
        }
        vars_obj = ClassificationVariables(obj)
        acts_obj = ClassificationActions(obj)

        score_rules = []
        decision_rules = []
        for rule in unified_rules:
            acts = rule.get("actions", []) or []
            if any(a.get("name") == "add_score" for a in acts):
                score_rules.append(rule)
            if any(a.get("name") == "set_classification" for a in acts):
                decision_rules.append(rule)

        run_all(
            rule_list=score_rules,
            defined_variables=vars_obj,
            defined_actions=acts_obj,
            stop_on_first_trigger=False,
        )

        def _is_high(rule):
            for a in rule.get("actions", []) or []:
                if a.get("name") == "set_classification":
                    rid = str(a.get("params", {}).get("rule_id", ""))
                    return rid.endswith("-H")
            return False

        decision_rules.sort(key=lambda r: (not _is_high(r)))

        run_all(
            rule_list=decision_rules,
            defined_variables=vars_obj,
            defined_actions=acts_obj,
            stop_on_first_trigger=False,
        )

        final_category = obj.get("category_path", "")
        level = obj.get("result_level", "")
        rid = obj.get("result_rule_id", "")
        audits = obj.get("audits", [])
        score = obj.get("score", 0)
        marker = obj.get("data_marker", "")
        hits = obj.get("hits", [])
        audit_str = ";".join([json.dumps(a, ensure_ascii=False) for a in audits]) if audits else ""
        tags_str = " ".join([str(h) for h in hits if h]) if hits else ""

        if final_category or level or rid:
            matched_rows += 1

        processed.append({"row": r, "category": final_category, "level": level, "rid": rid, "audit": audit_str, "score": score, "marker": marker, "tags": tags_str})

    max_depth = 0
    for p in processed:
        parts = [x for x in str(p["category"]).split("/") if x]
        if len(parts) > max_depth:
            max_depth = len(parts)

    cat_headers = [f"{i}级分类" for i in range(1, max_depth + 1)]
    new_headers = headers + cat_headers + ["数据标识", "分级", "规则ID", "命中标签", "置信度"]
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = ws.title
    out_ws.append(new_headers)

    for p in processed:
        parts = [x for x in str(p["category"]).split("/") if x]
        rid = str(p.get("rid", "") or "")
        if rid.endswith("-H"):
            conf = "high"
        elif rid.endswith("-M"):
            conf = "medium"
        else:
            conf = "low"

        if conf == "low":
            cat_cols = [""] * max_depth
            level = ""
            rid = ""
        else:
            cat_cols = parts + ([""] * (max_depth - len(parts)))
            level = p["level"]
            rid = p["rid"]

        out_ws.append([*p["row"], *cat_cols, p.get("marker", ""), level, rid, p.get("tags", ""), conf])

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    out_wb.save(out_path)
    print(f"[INFO] Saved result to: {out_path}")

    # 末尾调试统计信息
    ratio = (matched_rows / total_rows) if total_rows > 0 else 0.0
    print(f"[INFO] Stats: total_rows={total_rows}, matched_rows={matched_rows}, matched_ratio={ratio:.2%}")


def process_domain(
    domain: str,
    input_file: str,
    stop_first: bool,
    sheet_name: str,
):
    root = os.path.dirname(os.path.dirname(__file__))
    if input_file:
        base = os.path.splitext(os.path.basename(input_file))[0]
        out_dir = os.path.join(root, "outputs", domain)
        out_path = os.path.join(out_dir, base + ".classified.xlsx")
        classify_rows(
            domain, input_file, out_path, stop_first, sheet_name
        )
        print(out_path)
        return

    in_dir = os.path.join(root, "test", "data", domain)
    if not os.path.exists(in_dir):
        print(f"[ERROR] Input directory not found: {in_dir}")
        return

    files = [f for f in os.listdir(in_dir) if f.lower().endswith(".xlsx")]
    print(f"[DEBUG] Found files: {files}")
    for f in tqdm(files, desc=f"[FILES] {domain}", unit="file"):
        in_path = os.path.join(in_dir, f)
        base = os.path.splitext(f)[0]
        out_dir = os.path.join(root, "outputs", domain)
        out_path = os.path.join(out_dir, base + ".classified.xlsx")
        classify_rows(
            domain, in_path, out_path, stop_first, sheet_name
        )
        print(out_path)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("domain")
    parser.add_argument("--input", dest="input", default=None)
    parser.add_argument("--stop-first", dest="stop_first", default="false")
    parser.add_argument("--sheet", dest="sheet", default="")
    args = parser.parse_args()

    stop_first = str(args.stop_first).lower() != "false"
    process_domain(args.domain, args.input, stop_first, args.sheet)


if __name__ == "__main__":
    main()
