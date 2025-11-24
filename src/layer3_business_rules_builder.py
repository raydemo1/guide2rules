import json
import os
from typing import Dict, List, Any
from business_rules import export_rule_data

from rules.variables import ClassificationVariables
from rules.actions import ClassificationActions


def read_json(path: str) -> Any:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def list_extractions(domain_dir: str) -> List[str]:
    files = []
    for f in os.listdir(domain_dir):
        if f.endswith(".extraction.json") or f.endswith(".extraction.detailed.json") or f.endswith(".items.normalized.json"):
            files.append(os.path.join(domain_dir, f))
    return files


def try_read_csv(path: str) -> List[Dict]:
    import csv
    with open(path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        return [{k: (v or "") for k, v in row.items()} for row in reader]


def try_read_xlsx(path: str) -> List[Dict]:
    try:
        import openpyxl
    except Exception:
        return []
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    out: List[Dict] = []
    for row in ws.iter_rows(min_row=2):
        values = [c.value for c in row]
        out.append({h: ("" if v is None else str(v)) for h, v in zip(headers, values)})
    return out


def read_dynamic_rules(domain: str, root: str) -> List[Dict]:
    dir_path = os.path.join(root, "excels", domain)
    if not os.path.isdir(dir_path):
        return []
    files = os.listdir(dir_path)
    rows: List[Dict] = []
    for f in files:
        p = os.path.join(dir_path, f)
        if f.lower().endswith(".csv"):
            rows.extend(try_read_csv(p))
        elif f.lower().endswith(".xlsx"):
            rows.extend(try_read_xlsx(p))
    normalized: List[Dict] = []
    for i, r in enumerate(rows, 1):
        normalized.append({
            "RuleId": f"D-{i}",
            "FieldName": r.get("FieldName", r.get("field", "")),
            "Category": r.get("Category", r.get("category", "")),
            "Level": r.get("Level", r.get("level", "")),
            "ConditionExpression": r.get("Pattern", r.get("condition", "")),
            "ExceptionExpression": r.get("Exception", r.get("exception", "")),
            "Citation": r.get("Rationale", r.get("citation", "")),
            "PatternKeywords": r.get("PatternKeywords", r.get("keywords", "")),
            "PatternRegex": r.get("PatternRegex", r.get("regex", "")),
            "Priority": int(r.get("Priority", r.get("priority", 60)) or 60),
            "Source": f
        })
    return normalized


def _norm(s: Any) -> str:
    if s is None:
        return ""
    return str(s).strip()


def _level_rank(level: str) -> int:
    m = {"S1": 1, "S2": 2, "S3": 3, "S4": 4}
    return m.get(_norm(level).upper(), 0)


def resolve_conflicts(rows: List[Dict]) -> List[Dict]:
    buckets: Dict[str, Dict] = {}
    for r in rows:
        k1 = _norm(r.get("FieldName")).lower()
        k2 = _norm(r.get("Category")).lower()
        if not k1 and not k2:
            continue
        key = k1 + "|" + k2
        cur = buckets.get(key)
        if not cur:
            buckets[key] = r
            continue
        a = cur
        b = r
        ra = _level_rank(a.get("Level"))
        rb = _level_rank(b.get("Level"))
        chosen = a
        other = b
        if rb > ra:
            chosen = b
            other = a
        elif rb == ra:
            try:
                pa = int(a.get("Priority", 0))
            except Exception:
                pa = 0
            try:
                pb = int(b.get("Priority", 0))
            except Exception:
                pb = 0
            if pb < pa:
                chosen = b
                other = a
        conds = []
        for s in [_norm(chosen.get("ConditionExpression")), _norm(other.get("ConditionExpression"))]:
            if s:
                conds.append(s)
        exs = []
        for s in [_norm(chosen.get("ExceptionExpression")), _norm(other.get("ExceptionExpression"))]:
            if s:
                exs.append(s)
        cits = []
        for s in [_norm(chosen.get("Citation")), _norm(other.get("Citation"))]:
            if s:
                cits.append(s)
        srcs = []
        for s in [_norm(chosen.get("Source")), _norm(other.get("Source"))]:
            if s:
                srcs.append(s)
        pkw = []
        for s in [_norm(chosen.get("PatternKeywords")), _norm(other.get("PatternKeywords"))]:
            if s:
                pkw.extend([x.strip() for x in s.split(",") if x.strip()])
        prx = []
        for s in [_norm(chosen.get("PatternRegex")), _norm(other.get("PatternRegex"))]:
            if s:
                prx.extend([x.strip() for x in s.split("||") if x.strip()])
        chosen["ConditionExpression"] = " OR ".join(sorted(set(conds))) if conds else ""
        chosen["ExceptionExpression"] = " OR ".join(sorted(set(exs))) if exs else ""
        chosen["Citation"] = " || ".join(sorted(set(cits))) if cits else ""
        chosen["Source"] = ";".join(sorted(set(srcs)))
        chosen["PatternKeywords"] = ",".join(sorted(set(pkw))) if pkw else ""
        chosen["PatternRegex"] = "||".join(sorted(set(prx))) if prx else ""
        buckets[key] = chosen
    merged = list(buckets.values())
    for i, r in enumerate(merged, 1):
        r["RuleId"] = f"R-{i}"
    return merged


def _build_rule_from_row(r: Dict) -> Dict:
    field = _norm(r.get("FieldName"))
    category = _norm(r.get("Category"))
    level = _norm(r.get("Level"))
    rid = _norm(r.get("RuleId"))
    citation = _norm(r.get("Citation"))
    source = _norm(r.get("Source"))
    kw = [x for x in (_norm(r.get("PatternKeywords")) or "").split(",") if x]
    rx = [x for x in (_norm(r.get("PatternRegex")) or "").split("||") if x]
    cond_all = [
        {"name": "field_name", "operator": "equal_to", "value": field},
        {"name": "category_path", "operator": "equal_to", "value": category},
    ]
    cond_any = []
    for k in kw:
        cond_any.append({"name": "value_text", "operator": "contains", "value": k})
    for p in rx:
        cond_any.append({"name": "value_text", "operator": "matches_regex", "value": p})
    conditions = {"all": cond_all}
    if cond_any:
        conditions["any"] = cond_any
    actions = [
        {"name": "set_classification", "params": {"level": level, "rule_id": rid}},
        {"name": "append_audit", "params": {"citation": citation, "source": source}},
    ]
    return {"conditions": conditions, "actions": actions}


def write_rules(domain: str, root: str, rules: List[Dict]):
    out_dir = os.path.join(root, "rules", domain)
    os.makedirs(out_dir, exist_ok=True)
    rules_path = os.path.join(out_dir, "rules.json")
    with open(rules_path, "w", encoding="utf-8") as f:
        json.dump(rules, f, ensure_ascii=False, indent=2)
    data = export_rule_data(ClassificationVariables, ClassificationActions)
    with open(os.path.join(out_dir, "export_rule_data.json"), "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return rules_path


def main():
    root = os.path.dirname(os.path.dirname(__file__))
    artifacts = os.path.join(root, "artifacts")
    excels = os.path.join(root, "excels")
    domains = set()
    if os.path.isdir(artifacts):
        domains.update([d for d in os.listdir(artifacts) if os.path.isdir(os.path.join(artifacts, d))])
    if os.path.isdir(excels):
        domains.update([d for d in os.listdir(excels) if os.path.isdir(os.path.join(excels, d))])
    for domain in sorted(domains):
        static_rows: List[Dict] = []
        domain_artifacts = os.path.join(artifacts, domain)
        if os.path.isdir(domain_artifacts):
            ex_files = list_extractions(domain_artifacts)
            for p in ex_files:
                ext = read_json(p)
                items = ext.get("extraction", [])
                for it in items:
                    if isinstance(it.get("items"), list):
                        pth = it.get("path", {})
                        citation = json.dumps(it.get("citation", {}), ensure_ascii=False)
                        category = "/".join([pth.get("level1",""), pth.get("level2",""), pth.get("level3",""), pth.get("level4","")]).strip("/")
                        for sub in it["items"]:
                            static_rows.append({
                                "RuleId": f"S-{len(static_rows)+1}",
                                "FieldName": sub.get("name") or sub.get("canonical", ""),
                                "Category": category,
                                "Level": sub.get("level", ""),
                                "ConditionExpression": ";".join(sub.get("conditions", []) or []),
                                "ExceptionExpression": ";".join(sub.get("exceptions", []) or []),
                                "Citation": citation,
                                "PatternKeywords": ",".join((sub.get("patterns", {}) or {}).get("keywords", []) or []),
                                "PatternRegex": "||".join((sub.get("patterns", {}) or {}).get("regex", []) or []),
                                "Priority": 50,
                                "Source": ext.get("source", "")
                            })
                    else:
                        field = it.get("item", {}).get("name") or it.get("item", {}).get("canonical") or it.get("field", "")
                        if it.get("path"):
                            pth = it.get("path")
                            category = "/".join([pth.get("level1",""), pth.get("level2",""), pth.get("level3",""), pth.get("level4","")]).strip("/")
                        else:
                            category = it.get("category", "")
                        patt = (it.get("item", {}) or {}).get("patterns", {})
                        static_rows.append({
                            "RuleId": f"S-{len(static_rows)+1}",
                            "FieldName": field,
                            "Category": category,
                            "Level": it.get("level", ""),
                            "ConditionExpression": ";".join(it.get("conditions", []) or []),
                            "ExceptionExpression": ";".join(it.get("exceptions", []) or []),
                            "Citation": json.dumps(it.get("citation", {}), ensure_ascii=False),
                            "PatternKeywords": ",".join(patt.get("keywords", []) or []),
                            "PatternRegex": "||".join(patt.get("regex", []) or []),
                            "Priority": 50,
                            "Source": ext.get("source", "")
                        })
        dynamic_rows = read_dynamic_rules(domain, root)
        combined = resolve_conflicts(static_rows + dynamic_rows)
        rules = [_build_rule_from_row(r) for r in combined]
        path = write_rules(domain, root, rules)
        print(path)


if __name__ == "__main__":
    main()

